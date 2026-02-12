#!/usr/bin/env   python3

from openpyxl import load_workbook
import dbSqliteAlpameter as db

name_list = []

def import_price_xls():
    db.deleta_all_produts()
    #return

    print('Main Begin')
    wb = load_workbook('Products.xlsx')

    for sheet_name in wb.sheetnames:
        print(sheet_name)
        if sheet_name is None:
            continue
        if sheet_name == 'SUMMARY' \
            or sheet_name == 'Contents' \
            or sheet_name == 'Packaging' \
            or sheet_name == 'Wear Tapes' \
            or sheet_name == 'Modifications' \
            or sheet_name == 'Tools | Holders' \
            or sheet_name == 'Springs' :
            continue
        sheet = wb[sheet_name]
        # to print the maximum number of occupied rows in console
        #print(sheet.max_row)
        # считываем значение определенной ячейки
        #val = sheet['A3'].value
        #print(val)
        # iterate over all rows
        for i in range(7, sheet.max_row):
           # new = str(sheet.cell(row=i, column=4).value)
            #print(new)
           # if new.find('New') == -1:
            #    continue
            part_number = sheet.cell(row=i, column=9).value


            if str(part_number).find('=') >= 0:
                part_number = sheet.cell(row=i, column=1).value
            #print(part_number)

            #part_number = sheet.cell(row=i, column=1).value


            if part_number is None:
                continue
            name_material = str(sheet.cell(row=i, column=3).value)
            if name_material.startswith("TFE"):
                name_material = 'TFEP | FEPM'
            #print(name_material)
            id_nom = sheet.cell(row=i, column=1).value

            od_nom = sheet.cell(row=i, column=2).value
            #print(od_nom)
            len = sheet.cell(row=i, column=3).value
            #print(len)
            weight = sheet.cell(row=i, column=5).value
            #print(weight)

            res = db.get_product_for_part_number(part_number)
            if res is None:
                print(f'add {part_number}')
                db.create_new_product(part_number, \
                                  name_material, \
                                  id_nom, od_nom, \
                                  len, weight )

def import_price_csv(fn):
    print('Main Begin')
    fileread = open(fn, 'r')
    countstr = 0
    name_material = ""
    for line in fileread:
        str_array = line.split(';')
        #        print(str_array)
        countstr += 1
        if countstr == 2:
            name_material = str_array[0]
            if db.get_materials_for_name(name_material) is None:
                db.create_new_material(name_material, "MAX_PUR", "")
                print(f'add to DB {name_material}')
                name_list.append(name_material)
        part_number = str_array[8]
#        print(part_number)
#        print(name_material)
        id_nom = str_array[0]
#        print(id_nom)
        od_nom = str_array[1]
#        print(od_nom)
        len = str_array[2]
#        print(len)
        weight = str_array[4]
#        print(weight)
        f_ok = False
        try:
            idd = int(part_number)
            f_ok = True
        except:
          #  print('err')
          #  print(str_array)
            continue

        # if id_nom == "60":
        #     pass
        #
        # if part_number == "7700600082155":
        #     pass

        pr = db.get_product_for_part_number(part_number)
        if pr is None:
            print(f'add {part_number}')
            db.create_new_product(part_number, \
                                  name_material, \
                                  id_nom, od_nom, \
                                  len, weight)
        else:
            pr.name_material = name_material
            pr.len = len
            pr.weight = weight
            pr.save()


    fileread.close()


def import_all_prices():
#    db.deleta_all_produts() #from materials

    from pathlib import Path
    path_price = Path("price/")

    for fn in path_price.glob('*.csv'):
        import_price_csv(fn)

    print(f"name_list {name_list}")


if __name__ == '__main__':
    import_all_prices()
 #   import_price_xls()


    '''
H-Pur | FDA
SH-Pur
H-Pur TQ
H-Pur Green
H-Pur Blue | FDA
H-Pur Nat. | FDA
LT-Pur
XH-Pur
XSH-Pur
NBR
FKM | FDA
FKM Black
EPDM | FDA
H-NBR
MVQ | FDA
MVQ Blue | FDA
MVQ Nat | FDA
TFEP | FEPM | Aflas
POM | FDA
PA6 | FDA
PTFE-1 | FDA
PTFE-2
PTFE-3
PTFE-4
    '''